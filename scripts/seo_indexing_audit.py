from __future__ import annotations

import argparse
import csv
import re
import time
import xml.etree.ElementTree as ElementTree  # noqa: S405
from collections import Counter, defaultdict, deque
from collections.abc import Iterable
from dataclasses import dataclass, field
from html.parser import HTMLParser
from pathlib import Path
from urllib.parse import parse_qs, urldefrag, urljoin, urlparse, urlunparse

import requests
from openpyxl import load_workbook

PARAMS = {"sort", "order", "rating", "limit", "category_id", "from_admin"}
NOINDEX_PARAMS = {"sort", "order", "rating", "from_admin"}
DUPLICATE_SUFFIX_RE = re.compile(r"(?:-copy(?:-\d+)?|-[123])/?$", re.I)
COPY_SUFFIX_RE = re.compile(r"-copy(?:-\d+)?(?=/$|$)", re.I)
NUM_SUFFIX_RE = re.compile(r"-[123](?=/$|$)", re.I)
STATUS_404_RE = re.compile(r"not found|404|soft 404", re.I)
CRAWLED_RE = re.compile(r"crawled.*currently not indexed|currently not indexed", re.I)
DUP_GSC_RE = re.compile(r"duplicate without user-selected canonical|duplicate", re.I)
INDEXED_RE = re.compile(r"indexed|submitted and indexed", re.I)


class PageParser(HTMLParser):
    def __init__(self) -> None:
        super().__init__()
        self.links: list[str] = []
        self.canonical = ""
        self.title = ""
        self.h1 = ""
        self.robots = ""
        self._tag = ""
        self._text: list[str] = []

    def handle_starttag(self, tag: str, attrs: list[tuple[str, str | None]]) -> None:
        attrs_d = {k.lower(): (v or "") for k, v in attrs}
        if tag == "a" and attrs_d.get("href"):
            self.links.append(attrs_d["href"])
        if tag == "link" and "canonical" in attrs_d.get("rel", "").lower():
            self.canonical = attrs_d.get("href", "")
        if tag == "meta" and attrs_d.get("name", "").lower() == "robots":
            self.robots = attrs_d.get("content", "")
        if tag in {"title", "h1"}:
            self._tag = tag

    def handle_endtag(self, tag: str) -> None:
        if tag == self._tag:
            self._tag = ""

    def handle_data(self, data: str) -> None:
        text = data.strip()
        if self._tag == "title":
            self.title += text
        if self._tag == "h1" and not self.h1:
            self.h1 = text
        self._text.append(data)

    @property
    def word_count(self) -> int:
        return len(re.findall(r"\w+", " ".join(self._text)))


@dataclass
class Row:
    url: str
    status: str = ""
    final_url: str = ""
    canonical: str = ""
    title: str = ""
    h1: str = ""
    robots: str = ""
    word_count: int = 0
    internal_links: int = 0
    redirect_chain: str = ""
    source: set[str] = field(default_factory=set)
    gsc_reasons: set[str] = field(default_factory=set)


def norm(url: str) -> str:
    url, _ = urldefrag(str(url or "").strip())
    p = urlparse(url)
    if not p.scheme or not p.netloc:
        return ""
    return urlunparse((p.scheme.lower(), p.netloc.lower(), p.path or "/", "", p.query, ""))


def same_host(url: str, host: str) -> bool:
    return urlparse(url).netloc.lower().removeprefix("www.") == host.removeprefix("www.")


def candidate_primary(url: str, known_good: set[str]) -> str:
    p = urlparse(url)
    queryless = urlunparse((p.scheme, p.netloc, p.path, "", "", ""))
    clean = NUM_SUFFIX_RE.sub("", COPY_SUFFIX_RE.sub("", p.path))
    candidates = []
    if clean != p.path:
        candidates.extend(
            [
                urlunparse((p.scheme, p.netloc, clean, "", "", "")),
                urlunparse((p.scheme, p.netloc, clean.rstrip("/") + "/", "", "", "")),
            ]
        )
    candidates.append(queryless)
    for c in candidates:
        if c in known_good:
            return c
    return candidates[0] if clean != p.path else queryless if queryless != url else ""


def read_csv_rows(path: Path) -> Iterable[dict[str, str]]:
    with path.open(newline="", encoding="utf-8-sig") as f:
        yield from csv.DictReader(f)


def infer_gsc_reason(path: Path, text_values: Iterable[str]) -> str:
    blob = " ".join([path.name, *[str(v) for v in text_values if v]]).lower()
    if "not found" in blob or "לא נמצא" in blob or "404" in blob:
        return "Not Found (404)"
    if "duplicate without user-selected canonical" in blob or "עותק משוכפל" in blob:
        return "Duplicate without user-selected canonical"
    if ("crawled" in blob and "currently not indexed" in blob) or "נסרק - לא נכלל באינדקס" in blob:
        return "Crawled - Currently Not Indexed"
    if "indexed" in blob:
        return "Indexed Pages"
    return ""


def xlsx_rows(path: Path) -> tuple[list[dict[str, str]], str]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    all_strings: list[str] = []
    parsed_rows: list[dict[str, str]] = []
    for sheet in workbook.worksheets:
        values = list(sheet.iter_rows(values_only=True))
        all_strings.extend(str(cell) for row in values for cell in row if cell is not None)
        header_index = None
        headers: list[str] = []
        for index, row in enumerate(values[:25]):
            normalized = [str(cell).strip() if cell is not None else "" for cell in row]
            if any(cell.lower() in {"url", "page", "address", "כתובת אתר"} for cell in normalized):
                header_index = index
                headers = normalized
                break
        if header_index is None:
            continue
        for row in values[header_index + 1 :]:
            item = {
                headers[i]: str(cell).strip() for i, cell in enumerate(row) if i < len(headers) and cell is not None
            }
            if any(norm(value) for value in item.values()):
                parsed_rows.append(item)
    return parsed_rows, infer_gsc_reason(path, all_strings)


def write_converted_csv(path: Path, rows: list[dict[str, str]]) -> None:
    if not rows:
        return
    fields = sorted({key for row in rows for key in row})
    converted = path.with_suffix(".converted.csv")
    with converted.open("w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=fields)
        writer.writeheader()
        writer.writerows(rows)


def first_value(row: dict[str, str], *names: str) -> str:
    lower = {k.strip().lower(): v for k, v in row.items() if k}
    for name in names:
        if name.lower() in lower and lower[name.lower()]:
            return lower[name.lower()].strip()
    return ""


def parse_sitemap_text(text: str) -> list[str]:
    root = ElementTree.fromstring(text)  # noqa: S314
    ns = {"sm": "http://www.sitemaps.org/schemas/sitemap/0.9"}
    urls = [x.text.strip() for x in root.findall(".//sm:url/sm:loc", ns) if x.text]
    maps = [x.text.strip() for x in root.findall(".//sm:sitemap/sm:loc", ns) if x.text]
    return urls or maps


def import_record(raw: dict[str, str], path: Path, rows: dict[str, Row], inferred_reason: str = "") -> None:
    url = norm(first_value(raw, "url", "address", "page", "landing page", "top pages", "כתובת אתר"))
    if not url:
        url = next((norm(value) for value in raw.values() if norm(value)), "")
    if not url:
        return
    row = rows.setdefault(url, Row(url=url))
    row.source.add(f"{path.suffix.lower().lstrip('.') or 'file'}:{path.name}")
    row.status = first_value(raw, "status code", "status", "http status code") or row.status
    row.canonical = norm(first_value(raw, "canonical link element 1", "canonical", "canonical url")) or row.canonical
    row.title = first_value(raw, "title 1", "title", "page title") or row.title
    row.h1 = first_value(raw, "h1-1", "h1 1", "h1") or row.h1
    row.robots = first_value(raw, "meta robots 1", "meta robots", "robots") or row.robots
    row.word_count = int(first_value(raw, "word count") or row.word_count or 0)
    row.internal_links = int(first_value(raw, "inlinks", "unique inlinks", "internal links") or row.internal_links or 0)
    reason = (
        first_value(raw, "reason", "coverage state", "indexing state", "issue", "verdict", "בעיה") or inferred_reason
    )
    if reason:
        row.gsc_reasons.add(reason)


def import_files(paths: list[Path], rows: dict[str, Row]) -> None:
    for path in paths:
        if path.suffix.lower() == ".xml":
            for url in parse_sitemap_text(path.read_text(encoding="utf-8")):
                u = norm(url)
                if u:
                    rows.setdefault(u, Row(url=u)).source.add(f"sitemap:{path.name}")
            continue
        if path.suffix.lower() in {".txt", ".list"}:
            for line in path.read_text(encoding="utf-8").splitlines():
                u = norm(line)
                if u:
                    rows.setdefault(u, Row(url=u)).source.add(f"url-list:{path.name}")
            continue
        if path.suffix.lower() == ".xlsx":
            xrows, inferred_reason = xlsx_rows(path)
            write_converted_csv(path, xrows)
            for raw in xrows:
                import_record(raw, path, rows, inferred_reason)
            continue
        if path.suffix.lower() != ".csv":
            continue
        inferred_reason = infer_gsc_reason(path, [value for raw in read_csv_rows(path) for value in raw.values()])
        for raw in read_csv_rows(path):
            import_record(raw, path, rows, inferred_reason)


def discover_sitemaps(root: str, session: requests.Session, timeout: int) -> list[str]:
    try:
        robots = session.get(urljoin(root, "/robots.txt"), timeout=timeout).text
    except requests.RequestException:
        return [urljoin(root, "/sitemap.xml")]
    maps = re.findall(r"(?im)^sitemap:\s*(\S+)", robots)
    defaults = ["/sitemap.xml", "/sitemap-products.xml", "/sitemap-categories.xml", "/sitemap-blog.xml"]
    return list(dict.fromkeys(maps + [urljoin(root, d) for d in defaults]))


def fetch_live(root: str, max_pages: int, timeout: int, delay: float, rows: dict[str, Row]) -> None:
    host = urlparse(root).netloc.lower().removeprefix("www.")
    s = requests.Session()
    s.headers.update({"User-Agent": "CompassSEOAudit/2.0 (+real indexing audit)"})
    for sm in discover_sitemaps(root, s, timeout):
        try:
            for u in parse_sitemap_text(s.get(sm, timeout=timeout).text):
                u = norm(u)
                if u and same_host(u, host):
                    rows.setdefault(u, Row(url=u)).source.add(f"live-sitemap:{sm}")
        except (ElementTree.ParseError, requests.RequestException):
            continue
    q = deque(rows or {norm(root): Row(url=norm(root))})
    seen: set[str] = set()
    inbound = Counter()
    while q and len(seen) < max_pages:
        url = q.popleft()
        if url in seen or not same_host(url, host):
            continue
        seen.add(url)
        try:
            r = s.get(url, timeout=timeout, allow_redirects=True)
        except requests.RequestException:
            continue
        row = rows.setdefault(url, Row(url=url))
        if not row.source:
            row.source.add("live-crawl")
        row.status = str(r.status_code)
        row.final_url = norm(r.url)
        row.redirect_chain = " > ".join([norm(x.url) for x in r.history] + [row.final_url])
        if "html" in r.headers.get("content-type", ""):
            p = PageParser()
            p.feed(r.text)
            row.canonical = norm(urljoin(row.final_url or url, p.canonical)) if p.canonical else row.canonical
            row.title = p.title or row.title
            row.h1 = p.h1 or row.h1
            row.robots = p.robots or row.robots
            row.word_count = p.word_count or row.word_count
            for href in p.links:
                nxt = norm(urljoin(row.final_url or url, href))
                if nxt and same_host(nxt, host):
                    inbound[nxt] += 1
                    if nxt not in seen and len(seen) + len(q) < max_pages:
                        q.append(nxt)
        time.sleep(delay)
    for url, count in inbound.items():
        if url in rows:
            rows[url].internal_links = max(rows[url].internal_links, count)


def write_csv(path: Path, fields: list[str], rows: list[dict[str, object]]) -> None:
    if not rows:
        if path.exists():
            path.unlink()
        return
    with path.open("w", newline="", encoding="utf-8-sig") as f:
        w = csv.DictWriter(f, fieldnames=fields)
        w.writeheader()
        w.writerows(rows)


def page_type(url: str) -> str:
    p = urlparse(url).path.lower()
    if "search" in p or "search" in urlparse(url).query.lower():
        return "search"
    if "sitemap" in p:
        return "sitemap"
    if "blog" in p or "article" in p:
        return "article"
    if "categor" in p or "category" in p:
        return "category"
    return "product"


def analyze(rows: dict[str, Row], out: Path) -> None:
    out.mkdir(parents=True, exist_ok=True)
    good = {u for u, r in rows.items() if r.status.startswith("2") or not r.status}
    redirects = []
    canon_rules = []
    noindex = []
    dups = []
    broken = []
    sitemap_issues = []
    nonindexed = []
    fixes = []
    summary = defaultdict(int)
    overlap = []
    for u, r in rows.items():
        qs = parse_qs(urlparse(u).query)
        has_param = bool(PARAMS & set(qs))
        is_dup = bool(DUPLICATE_SUFFIX_RE.search(urlparse(u).path))
        is_broken = r.status in {"404", "410"} or STATUS_404_RE.search(" ".join(r.gsc_reasons))
        is_redirect = r.status in {"301", "302"} or bool(r.redirect_chain and r.final_url and r.final_url != u)
        target = candidate_primary(u, good)
        if r.status.startswith("2") or not r.status:
            summary["valid"] += 1
        if is_broken:
            summary["404"] += 1
        if is_dup:
            summary["duplicates"] += 1
        if "-copy" in urlparse(u).path:
            summary["copy"] += 1
        if NUM_SUFFIX_RE.search(urlparse(u).path):
            summary["numeric_suffix"] += 1
        if is_redirect:
            summary["redirect"] += 1
        if has_param:
            summary["parameters"] += 1
        if {"sort", "order", "rating"} & set(qs):
            summary["sort_order_rating"] += 1
        if "category_id" in qs:
            summary["category_id"] += 1
        if re.search(r"-[1](?=/$|$)", urlparse(u).path):
            summary["suffix_1"] += 1
        if page_type(u) == "sitemap":
            summary["sitemap_urls"] += 1
        if page_type(u) == "search":
            summary["search_urls"] += 1
        if has_param:
            if NOINDEX_PARAMS & set(qs):
                noindex.append({"pattern": u})
                action = "Noindex"
                summary["needs_noindex"] += 1
            else:
                canon_rules.append({"pattern": u, "canonical_target": target})
                action = "Canonical"
                summary["needs_canonical"] += 1
            fixes.append({"url": u, "issue_type": "parameters", "suggested_action": action, "target_url": target})
        if is_dup:
            dups.append({"url": u, "duplicate_of": target})
            if target:
                redirects.append({"old_url": u, "new_url": target})
                summary["needs_redirect"] += 1
            fixes.append(
                {"url": u, "issue_type": "duplicate", "suggested_action": "Redirect 301", "target_url": target}
            )
        if is_broken:
            action = f"Redirect 301 to {target}" if target else "Remove internal links / restore page"
            broken.append({"url": u, "recommended_action": action})
            fixes.append({"url": u, "issue_type": "404/410/soft 404", "suggested_action": action, "target_url": target})
        reasons = " | ".join(sorted(r.gsc_reasons))
        if page_type(u) == "sitemap" and reasons:
            sitemap_issues.append({"url": u, "issue": reasons})
        if is_redirect and any(s.startswith("sitemap") or s.startswith("live-sitemap") for s in r.source):
            sitemap_issues.append({"url": u, "issue": "sitemap URL redirects"})
        if "noindex" in r.robots.lower() and any(
            s.startswith("sitemap") or s.startswith("live-sitemap") for s in r.source
        ):
            sitemap_issues.append({"url": u, "issue": "sitemap URL is noindex"})
        if (
            r.canonical
            and r.canonical != u
            and any(s.startswith("sitemap") or s.startswith("live-sitemap") for s in r.source)
        ):
            sitemap_issues.append({"url": u, "issue": f"canonical points to {r.canonical}"})
        if re.search(r"crawled|not indexed|discovered", reasons, re.I):
            content_type = page_type(u)
            if content_type == "product":
                summary["nonindexed_products"] += 1
            if content_type == "article":
                summary["nonindexed_articles"] += 1
            nonindexed.append(
                {
                    "url": u,
                    "type": content_type,
                    "reason": reasons,
                    "recommendation": "Improve content depth, canonical clarity and internal links",
                }
            )
        if len({x for x in [has_param, is_dup, is_broken, bool(nonindexed)] if x}) > 1:
            overlap.append(u)
    write_csv(out / "redirects.csv", ["old_url", "new_url"], redirects)
    write_csv(out / "canonical-rules.csv", ["pattern", "canonical_target"], canon_rules)
    write_csv(out / "noindex-rules.csv", ["pattern"], noindex)
    write_csv(out / "duplicate-urls.csv", ["url", "duplicate_of"], dups)
    write_csv(out / "404-urls.csv", ["url", "recommended_action"], broken)
    write_csv(out / "sitemap-issues.csv", ["url", "issue"], sitemap_issues)
    write_csv(out / "non-indexed-content.csv", ["url", "type", "reason", "recommendation"], nonindexed)
    write_csv(out / "fix-recommendations.csv", ["url", "issue_type", "suggested_action", "target_url"], fixes)
    (out / "implementation-plan.md").write_text(
        "\n".join(
            [
                "# Compass Grill Real Indexing Audit Execution Plan",
                "",
                f"Analyzed real imported/live URLs: {len(rows)}.",
                "",
                "## Numeric summary",
                *[
                    f"- {k}: {summary[k]}"
                    for k in [
                        "valid",
                        "404",
                        "duplicates",
                        "copy",
                        "numeric_suffix",
                        "redirect",
                        "parameters",
                        "needs_redirect",
                        "needs_canonical",
                        "needs_noindex",
                        "sort_order_rating",
                        "category_id",
                        "suffix_1",
                        "sitemap_urls",
                        "search_urls",
                        "nonindexed_products",
                        "nonindexed_articles",
                    ]
                ],
                "",
                "## Priority order",
                "1. 404 URLs with internal links or GSC impressions.",
                "2. 404 URLs also present in duplicate/non-indexed reports.",
                "3. Duplicate copy/numeric suffix URLs.",
                "4. Parameter URLs requiring noindex/canonical.",
                "5. Non-indexed articles/categories/products needing content and links.",
                "",
                f"High-overlap URLs requiring manual review first: {len(overlap)}.",
            ]
        ),
        encoding="utf-8",
    )


def main() -> int:
    ap = argparse.ArgumentParser(description="Real Compass Grill indexing audit from live crawl and/or exports")
    ap.add_argument("--root", default="https://compassgrill.co.il/")
    ap.add_argument("--out", default="reports/compassgrill-indexing-audit")
    ap.add_argument("--input", action="append", default=[], help="CSV/XML/TXT export file; repeatable")
    ap.add_argument("--input-dir", help="Directory containing GSC, sitemap, Screaming Frog, or URL-list exports")
    ap.add_argument("--live", action="store_true", help="Attempt live crawl in addition to imports")
    ap.add_argument("--max-pages", type=int, default=5000)
    ap.add_argument("--timeout", type=int, default=20)
    ap.add_argument("--delay", type=float, default=0.05)
    args = ap.parse_args()
    rows: dict[str, Row] = {}
    inputs = [Path(x) for x in args.input]
    if args.input_dir:
        inputs.extend(p for p in Path(args.input_dir).iterdir() if p.is_file())
    if not inputs and not args.live:
        inputs.extend(sorted(Path.cwd().glob("*Coverage-Drilldown-*.converted.csv")))
        inputs.extend(sorted(Path.cwd().glob("*Coverage-Drilldown-*.xlsx")))
    if inputs:
        import_files(inputs, rows)
    if args.live:
        fetch_live(args.root, args.max_pages, args.timeout, args.delay, rows)
    if not rows:
        raise SystemExit(
            "No real URLs loaded. Provide --input/--input-dir exports or use --live from a network with access."
        )
    analyze(rows, Path(args.out))
    print(f"Analyzed {len(rows)} real URLs into {args.out}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
